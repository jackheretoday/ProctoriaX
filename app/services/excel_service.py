"""
Excel Service
Generates and encrypts Excel files for results export
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from app.services.encryption_service import EncryptionService
from app.utils.exceptions import ValidationError, DatabaseError
import os
import tempfile
import shutil
from datetime import datetime


class ExcelService:
    """Service for Excel file generation and encryption"""
    
    @staticmethod
    def generate_results_excel(test_id, test_name, results_data):
        """
        Generate Excel file with test results
        
        Args:
            test_id: ID of the test
            test_name: Name of the test
            results_data: List of result dictionaries
            
        Returns:
            str: Path to Excel file
            
        Raises:
            ValidationError: If validation fails
            DatabaseError: If generation fails
        """
        try:
            if not results_data:
                raise ValidationError("No results data to export")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            
            # Add header section with logo and branding (rows 1-4)
            # Row 1: XIE Logo image - centered and properly sized in C1
            try:
                # Get the path to the logo image
                logo_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'images', 'xie-logo.png')
                if os.path.exists(logo_path):
                    # Add the image to the Excel sheet
                    img = Image(logo_path)
                    # Set optimal image size for professional appearance
                    img.width = 180  # Medium size - not too big, not too small
                    img.height = 50  # Proportional height
                    
                    # Position the image in C1 cell for better centering
                    img.anchor = 'C1'  # Position in C1 cell
                    
                    # Add the image to the worksheet
                    ws.add_image(img)
                    
                    # Adjust row height for better visual balance
                    ws.row_dimensions[1].height = 50
                    
                    # Merge cells for proper alignment area (C1:E1)
                    ws.merge_cells('C1:E1')
                    
                    # Add some padding by adjusting column widths for header
                    ws.column_dimensions['A'].width = 5   # Smaller width for A column
                    ws.column_dimensions['B'].width = 5   # Smaller width for B column
                    ws.column_dimensions['C'].width = 20  # Logo column
                    ws.column_dimensions['D'].width = 15
                    ws.column_dimensions['E'].width = 15
                    
                    # The image will be centered within the merged area C1:E1
                else:
                    # Fallback to text if image not found
                    ws.merge_cells('C1:E1')
                    logo_cell = ws['C1']
                    logo_cell.value = "XIE LOGO"
                    logo_cell.font = Font(size=16, bold=True, color='4472C4')
                    logo_cell.alignment = Alignment(horizontal='center', vertical='center')
            except Exception as e:
                # Fallback to text if image loading fails
                ws.merge_cells('C1:E1')
                logo_cell = ws['C1']
                logo_cell.value = "XIE LOGO"
                logo_cell.font = Font(size=16, bold=True, color='4472C4')
                logo_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Row 2: Institute Name - centered with proper formatting
            ws.merge_cells('C2:E2')
            institute_cell = ws['C2']
            institute_cell.value = "Xavier Institute of Engineering"
            institute_cell.font = Font(size=14, bold=True, color='2F5597')
            institute_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Row 3: Powered by Proctoria - centered with subtle styling
            ws.merge_cells('C3:E3')
            powered_cell = ws['C3']
            powered_cell.value = "Powered by Proctoria"
            powered_cell.font = Font(size=11, italic=True, color='666666')
            powered_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Row 4: Empty row for spacing - ensure proper separation
            ws.merge_cells('C4:E4')
            ws.row_dimensions[4].height = 15  # Add some spacing
            
            # Row 5: Test Title
            title_row = 5
            ws.merge_cells(f'A{title_row}:G{title_row}')
            title_cell = ws[f'A{title_row}']
            title_cell.value = f"Test Results: {test_name}"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            title_cell.font = Font(size=14, bold=True, color='FFFFFF')
            
            # Add headers (next row after title)
            header_row = title_row + 1
            headers = ['Student Name', 'Roll Number', 'Total Questions', 'Correct Answers', 
                      'Wrong Answers', 'Score', 'Percentage']
            
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data rows
            for row_num, result in enumerate(results_data, start=header_row + 1):
                ws.cell(row=row_num, column=1).value = result.get('student_name', 'Unknown')
                ws.cell(row=row_num, column=2).value = result.get('roll_number', 'N/A')
                ws.cell(row=row_num, column=3).value = result.get('total_questions', 0)
                ws.cell(row=row_num, column=4).value = result.get('correct_answers', 0)
                ws.cell(row=row_num, column=5).value = result.get('wrong_answers', 0)
                ws.cell(row=row_num, column=6).value = result.get('score', 0)
                
                # Format percentage
                percentage_cell = ws.cell(row=row_num, column=7)
                percentage = result.get('percentage', 0)
                percentage_cell.value = f"{percentage}%"
                
                # Color code based on percentage
                if percentage >= 80:
                    percentage_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif percentage >= 50:
                    percentage_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    percentage_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # Auto-adjust column widths
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0
                
                for cell in ws[column_letter]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = max(max_length + 2, 10)  # Minimum width of 10
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary statistics at the bottom
            last_row = header_row + len(results_data)
            ws.cell(row=last_row + 2, column=1).value = "Aggregate Results:"
            ws.cell(row=last_row + 2, column=1).font = Font(bold=True, size=12)
            ws.cell(row=last_row + 2, column=1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            
            # Calculate statistics
            if results_data:
                total_students = len(results_data)
                avg_percentage = sum(r.get('percentage', 0) for r in results_data) / total_students
                max_percentage = max(r.get('percentage', 0) for r in results_data)
                min_percentage = min(r.get('percentage', 0) for r in results_data)
                avg_score = sum(r.get('score', 0) for r in results_data) / total_students
                
                ws.cell(row=last_row + 3, column=1).value = "Total Students:"
                ws.cell(row=last_row + 3, column=2).value = total_students
                ws.cell(row=last_row + 3, column=2).font = Font(bold=True)
                
                ws.cell(row=last_row + 4, column=1).value = "Average Score:"
                ws.cell(row=last_row + 4, column=2).value = f"{avg_score:.2f}"
                ws.cell(row=last_row + 4, column=2).font = Font(bold=True)
                
                ws.cell(row=last_row + 5, column=1).value = "Average Percentage:"
                ws.cell(row=last_row + 5, column=2).value = f"{avg_percentage:.2f}%"
                ws.cell(row=last_row + 5, column=2).font = Font(bold=True)
                
                ws.cell(row=last_row + 6, column=1).value = "Highest Percentage:"
                ws.cell(row=last_row + 6, column=2).value = f"{max_percentage}%"
                ws.cell(row=last_row + 6, column=2).font = Font(bold=True)
                
                ws.cell(row=last_row + 7, column=1).value = "Lowest Percentage:"
                ws.cell(row=last_row + 7, column=2).value = f"{min_percentage}%"
                ws.cell(row=last_row + 7, column=2).font = Font(bold=True)
                
                # Add grade distribution
                ws.cell(row=last_row + 9, column=1).value = "Grade Distribution:"
                ws.cell(row=last_row + 9, column=1).font = Font(bold=True, size=12)
                ws.cell(row=last_row + 9, column=1).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                
                excellent = sum(1 for r in results_data if r.get('percentage', 0) >= 80)
                good = sum(1 for r in results_data if 60 <= r.get('percentage', 0) < 80)
                average = sum(1 for r in results_data if 40 <= r.get('percentage', 0) < 60)
                poor = sum(1 for r in results_data if r.get('percentage', 0) < 40)
                
                ws.cell(row=last_row + 10, column=1).value = "Excellent (80%+):"
                ws.cell(row=last_row + 10, column=2).value = f"{excellent} students ({(excellent/total_students)*100:.1f}%)"
                ws.cell(row=last_row + 10, column=2).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                
                ws.cell(row=last_row + 11, column=1).value = "Good (60-79%):"
                ws.cell(row=last_row + 11, column=2).value = f"{good} students ({(good/total_students)*100:.1f}%)"
                ws.cell(row=last_row + 11, column=2).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                
                ws.cell(row=last_row + 12, column=1).value = "Average (40-59%):"
                ws.cell(row=last_row + 12, column=2).value = f"{average} students ({(average/total_students)*100:.1f}%)"
                ws.cell(row=last_row + 12, column=2).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                
                ws.cell(row=last_row + 13, column=1).value = "Poor (<40%):"
                ws.cell(row=last_row + 13, column=2).value = f"{poor} students ({(poor/total_students)*100:.1f}%)"
                ws.cell(row=last_row + 13, column=2).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # Save to temporary file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            temp_dir = tempfile.mkdtemp()
            excel_filename = f"test_{test_id}_results_{timestamp}.xlsx"
            temp_excel_path = os.path.join(temp_dir, excel_filename)
            
            wb.save(temp_excel_path)
            
            # Return the direct path
            return temp_excel_path
            
        except ValidationError:
            raise
        except Exception as e:
            # Clean up on error
            if 'temp_dir' in locals() and os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            raise DatabaseError(f"Error generating Excel file: {str(e)}")
    
    @staticmethod
    def decrypt_and_prepare_download(encrypted_file_path):
        """
        Decrypt Excel file and prepare for download
        
        Args:
            encrypted_file_path: Path to encrypted file
            
        Returns:
            str: Path to decrypted file
            
        Raises:
            ValidationError: If file not found
            DatabaseError: If decryption fails
        """
        try:
            if not os.path.exists(encrypted_file_path):
                raise ValidationError("Encrypted file not found")
            
            # Create temporary directory for decrypted file
            temp_dir = tempfile.mkdtemp()
            
            # Extract original filename (remove .enc extension)
            encrypted_filename = os.path.basename(encrypted_file_path)
            if encrypted_filename.endswith('.enc'):
                decrypted_filename = encrypted_filename[:-4]  # Remove .enc
            else:
                decrypted_filename = encrypted_filename + '.xlsx'
            
            decrypted_path = os.path.join(temp_dir, decrypted_filename)
            
            # Decrypt file
            EncryptionService.decrypt_file(encrypted_file_path, decrypted_path)
            
            return decrypted_path
            
        except ValidationError:
            raise
        except Exception as e:
            # Clean up on error
            if 'temp_dir' in locals() and os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            raise DatabaseError(f"Error decrypting Excel file: {str(e)}")
    
    @staticmethod
    def cleanup_temp_file(file_path):
        """
        Clean up temporary decrypted file
        
        Args:
            file_path: Path to temporary file
        """
        try:
            if file_path and os.path.exists(file_path):
                # Delete the file
                os.remove(file_path)
                
                # Delete the temporary directory if empty
                temp_dir = os.path.dirname(file_path)
                if os.path.exists(temp_dir) and not os.listdir(temp_dir):
                    os.rmdir(temp_dir)
        except:
            pass  # Ignore errors during cleanup
    
    @staticmethod
    def generate_detailed_results_excel(test_id, test_name, results_data, questions_data=None, teacher_name=None):
        """
        Generate detailed Excel with question-wise breakdown (optional enhancement)
        
        Args:
            test_id: ID of the test
            test_name: Name of the test
            results_data: List of result dictionaries
            questions_data: Optional question-wise answers
            
        Returns:
            str: Path to encrypted Excel file
        """
        try:
            # Create workbook with multiple sheets
            wb = Workbook()
            
            # Sheet 1: Summary
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Add summary content (same as generate_results_excel)
            # ... (implementation similar to above)
            
            # Sheet 2: Detailed (if questions_data provided)
            if questions_data:
                ws_detailed = wb.create_sheet("Detailed")
                # Add question-wise breakdown
                # ... (can be implemented later if needed)
            
            # Save and encrypt (same process as above)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            temp_dir = tempfile.mkdtemp()
            excel_filename = f"test_{test_id}_detailed_{timestamp}.xlsx"
            temp_excel_path = os.path.join(temp_dir, excel_filename)
            
            wb.save(temp_excel_path)
            
            # Encrypt
            encrypted_dir = os.path.join('storage', 'encrypted', 'results')
            os.makedirs(encrypted_dir, exist_ok=True)
            
            encrypted_filename = f"test_{test_id}_detailed_{timestamp}.enc"
            encrypted_path = os.path.join(encrypted_dir, encrypted_filename)
            
            EncryptionService.encrypt_file(temp_excel_path, encrypted_path)
            
            # Clean up
            shutil.rmtree(temp_dir)
            
            return encrypted_path
            
        except Exception as e:
            if 'temp_dir' in locals() and os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            raise DatabaseError(f"Error generating detailed Excel: {str(e)}")
